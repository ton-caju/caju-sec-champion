#!/usr/bin/env python3
"""
Script para gerar planilha OWASP Risk Rating
Desafio Caju Security Champions - Recuperação de Senha PID
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def criar_planilha_owasp():
    """Cria planilha OWASP Risk Rating completa"""

    wb = Workbook()

    # Criar abas
    ws_resumo = wb.active
    ws_resumo.title = "Resumo Executivo"
    ws_likelihood = wb.create_sheet("Likelihood Factors")
    ws_impact = wb.create_sheet("Impact Factors")
    ws_riscos = wb.create_sheet("Análise de Riscos")
    ws_mitigacoes = wb.create_sheet("Plano de Mitigação")

    # Estilos
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)

    critico_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    critico_font = Font(bold=True, color="FFFFFF")

    alto_fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
    alto_font = Font(bold=True, color="FFFFFF")

    medio_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    medio_font = Font(bold=True, color="000000")

    baixo_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    baixo_font = Font(bold=True, color="000000")

    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ============================================================
    # ABA 1: RESUMO EXECUTIVO
    # ============================================================

    ws_resumo['A1'] = "OWASP RISK RATING METHODOLOGY"
    ws_resumo['A1'].font = Font(bold=True, size=16, color="1F4E78")
    ws_resumo.merge_cells('A1:G1')

    ws_resumo['A2'] = "Desafio Caju Security Champions - Recuperação de Senha com Validação de Identidade (PID)"
    ws_resumo['A2'].font = Font(size=12, italic=True)
    ws_resumo.merge_cells('A2:G2')

    ws_resumo['A3'] = "Data: 2025-11-14"
    ws_resumo['A4'] = "Versão: 1.0"
    ws_resumo['A5'] = "Escopo: Frontend (React) + BFF (Node.js)"

    ws_resumo.row_dimensions[7].height = 25
    headers_resumo = ['Criticidade', 'Quantidade', '% Total', 'Descrição']
    for col, header in enumerate(headers_resumo, start=1):
        cell = ws_resumo.cell(row=7, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border

    # Dados do resumo
    resumo_data = [
        ('CRÍTICA', 5, '38%', 'Ameaças que permitem comprometimento imediato de contas/sistema'),
        ('ALTA', 5, '38%', 'Ameaças que facilitam significativamente ataques bem-sucedidos'),
        ('MÉDIA', 3, '24%', 'Ameaças que aumentam a superfície de ataque mas requerem outros vetores'),
        ('TOTAL', 13, '100%', 'Total de ameaças identificadas no fluxo de recuperação PID')
    ]

    for row_idx, (criticidade, qtd, percent, desc) in enumerate(resumo_data, start=8):
        ws_resumo.cell(row=row_idx, column=1).value = criticidade
        ws_resumo.cell(row=row_idx, column=2).value = qtd
        ws_resumo.cell(row=row_idx, column=3).value = percent
        ws_resumo.cell(row=row_idx, column=4).value = desc

        for col in range(1, 5):
            cell = ws_resumo.cell(row=row_idx, column=col)
            cell.border = thin_border
            cell.alignment = center_alignment if col <= 3 else left_alignment

            if criticidade == 'CRÍTICA':
                cell.fill = critico_fill
                cell.font = critico_font
            elif criticidade == 'ALTA':
                cell.fill = alto_fill
                cell.font = alto_font
            elif criticidade == 'MÉDIA':
                cell.fill = medio_fill
                cell.font = medio_font
            elif criticidade == 'TOTAL':
                cell.font = Font(bold=True)

    # Ajustar larguras
    ws_resumo.column_dimensions['A'].width = 15
    ws_resumo.column_dimensions['B'].width = 12
    ws_resumo.column_dimensions['C'].width = 10
    ws_resumo.column_dimensions['D'].width = 70

    # ============================================================
    # ABA 2: LIKELIHOOD FACTORS (Fatores de Probabilidade)
    # ============================================================

    ws_likelihood['A1'] = "LIKELIHOOD FACTORS (Fatores de Probabilidade)"
    ws_likelihood['A1'].font = Font(bold=True, size=14, color="1F4E78")
    ws_likelihood.merge_cells('A1:D1')

    ws_likelihood['A3'] = "OWASP divide Likelihood em fatores relacionados ao Atacante (Threat Agent) e à Vulnerabilidade:"
    ws_likelihood.merge_cells('A3:D3')

    # Tabela Threat Agent Factors
    ws_likelihood['A5'] = "THREAT AGENT FACTORS (Fatores do Atacante)"
    ws_likelihood['A5'].font = Font(bold=True, size=12)
    ws_likelihood.merge_cells('A5:D5')

    headers_threat = ['Fator', 'Baixo (0-2)', 'Médio (3-5)', 'Alto (6-9)']
    for col, header in enumerate(headers_threat, start=1):
        cell = ws_likelihood.cell(row=6, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border

    threat_agent_data = [
        ('Skill Level\n(Nível de Habilidade)',
         'Nenhuma habilidade técnica',
         'Alguma habilidade técnica',
         'Hacker profissional'),

        ('Motive\n(Motivação)',
         'Recompensa baixa',
         'Recompensa possível',
         'Recompensa alta'),

        ('Opportunity\n(Oportunidade)',
         'Acesso completo ou recursos caros necessários',
         'Acesso especial ou recursos necessários',
         'Algum acesso ou recursos'),

        ('Size\n(Tamanho)',
         'Desenvolvedores / Administradores',
         'Usuários autenticados',
         'Usuários anônimos na Internet')
    ]

    for row_idx, (fator, baixo, medio, alto) in enumerate(threat_agent_data, start=7):
        ws_likelihood.cell(row=row_idx, column=1).value = fator
        ws_likelihood.cell(row=row_idx, column=2).value = baixo
        ws_likelihood.cell(row=row_idx, column=3).value = medio
        ws_likelihood.cell(row=row_idx, column=4).value = alto

        for col in range(1, 5):
            cell = ws_likelihood.cell(row=row_idx, column=col)
            cell.border = thin_border
            cell.alignment = left_alignment
            ws_likelihood.row_dimensions[row_idx].height = 40

    # Tabela Vulnerability Factors
    ws_likelihood['A12'] = "VULNERABILITY FACTORS (Fatores da Vulnerabilidade)"
    ws_likelihood['A12'].font = Font(bold=True, size=12)
    ws_likelihood.merge_cells('A12:D12')

    for col, header in enumerate(headers_threat, start=1):
        cell = ws_likelihood.cell(row=13, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border

    vuln_factors_data = [
        ('Ease of Discovery\n(Facilidade de Descoberta)',
         'Praticamente impossível',
         'Difícil',
         'Fácil / Ferramentas automatizadas disponíveis'),

        ('Ease of Exploit\n(Facilidade de Exploração)',
         'Teórico',
         'Difícil',
         'Fácil / Exploits públicos disponíveis'),

        ('Awareness\n(Conscientização)',
         'Desconhecida',
         'Hidden (Obscura)',
         'Óbvia / Conhecida publicamente'),

        ('Intrusion Detection\n(Detecção de Intrusão)',
         'Ativa detecção em aplicação',
         'Logged e revisado',
         'Não logged')
    ]

    for row_idx, (fator, baixo, medio, alto) in enumerate(vuln_factors_data, start=14):
        ws_likelihood.cell(row=row_idx, column=1).value = fator
        ws_likelihood.cell(row=row_idx, column=2).value = baixo
        ws_likelihood.cell(row=row_idx, column=3).value = medio
        ws_likelihood.cell(row=row_idx, column=4).value = alto

        for col in range(1, 5):
            cell = ws_likelihood.cell(row=row_idx, column=col)
            cell.border = thin_border
            cell.alignment = left_alignment
            ws_likelihood.row_dimensions[row_idx].height = 40

    # Ajustar larguras
    ws_likelihood.column_dimensions['A'].width = 25
    ws_likelihood.column_dimensions['B'].width = 35
    ws_likelihood.column_dimensions['C'].width = 35
    ws_likelihood.column_dimensions['D'].width = 35

    # ============================================================
    # ABA 3: IMPACT FACTORS (Fatores de Impacto)
    # ============================================================

    ws_impact['A1'] = "IMPACT FACTORS (Fatores de Impacto)"
    ws_impact['A1'].font = Font(bold=True, size=14, color="1F4E78")
    ws_impact.merge_cells('A1:D1')

    ws_impact['A3'] = "OWASP divide Impact em Impacto Técnico e Impacto ao Negócio:"
    ws_impact.merge_cells('A3:D3')

    # Tabela Technical Impact
    ws_impact['A5'] = "TECHNICAL IMPACT (Impacto Técnico)"
    ws_impact['A5'].font = Font(bold=True, size=12)
    ws_impact.merge_cells('A5:D5')

    headers_impact = ['Fator', 'Baixo (0-2)', 'Médio (3-5)', 'Alto (6-9)']
    for col, header in enumerate(headers_impact, start=1):
        cell = ws_impact.cell(row=6, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border

    technical_impact_data = [
        ('Loss of Confidentiality\n(Perda de Confidencialidade)',
         'Dados não-sensíveis mínimos',
         'Dados não-sensíveis extensos ou sensíveis mínimos',
         'Dados sensíveis extensos'),

        ('Loss of Integrity\n(Perda de Integridade)',
         'Dados mínimos corrompidos',
         'Dados mínimos críticos ou extensos não-críticos corrompidos',
         'Dados críticos extensos corrompidos'),

        ('Loss of Availability\n(Perda de Disponibilidade)',
         'Serviços secundários mínimos interrompidos',
         'Serviços primários mínimos ou secundários extensos interrompidos',
         'Serviços primários extensos interrompidos'),

        ('Loss of Accountability\n(Perda de Rastreabilidade)',
         'Totalmente rastreável',
         'Possivelmente rastreável',
         'Completamente anônimo')
    ]

    for row_idx, (fator, baixo, medio, alto) in enumerate(technical_impact_data, start=7):
        ws_impact.cell(row=row_idx, column=1).value = fator
        ws_impact.cell(row=row_idx, column=2).value = baixo
        ws_impact.cell(row=row_idx, column=3).value = medio
        ws_impact.cell(row=row_idx, column=4).value = alto

        for col in range(1, 5):
            cell = ws_impact.cell(row=row_idx, column=col)
            cell.border = thin_border
            cell.alignment = left_alignment
            ws_impact.row_dimensions[row_idx].height = 40

    # Tabela Business Impact
    ws_impact['A12'] = "BUSINESS IMPACT (Impacto ao Negócio)"
    ws_impact['A12'].font = Font(bold=True, size=12)
    ws_impact.merge_cells('A12:D12')

    for col, header in enumerate(headers_impact, start=1):
        cell = ws_impact.cell(row=13, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border

    business_impact_data = [
        ('Financial Damage\n(Dano Financeiro)',
         'Menos que custo de correção',
         'Perda menor de receita',
         'Falência'),

        ('Reputation Damage\n(Dano à Reputação)',
         'Perda mínima',
         'Perda de contas importantes',
         'Perda de goodwill / marca'),

        ('Non-Compliance\n(Não-Conformidade)',
         'Violação menor',
         'Violação clara',
         'Violação de alto perfil'),

        ('Privacy Violation\n(Violação de Privacidade)',
         'Um indivíduo',
         'Centenas de pessoas',
         'Milhares de pessoas / LGPD')
    ]

    for row_idx, (fator, baixo, medio, alto) in enumerate(business_impact_data, start=14):
        ws_impact.cell(row=row_idx, column=1).value = fator
        ws_impact.cell(row=row_idx, column=2).value = baixo
        ws_impact.cell(row=row_idx, column=3).value = medio
        ws_impact.cell(row=row_idx, column=4).value = alto

        for col in range(1, 5):
            cell = ws_impact.cell(row=row_idx, column=col)
            cell.border = thin_border
            cell.alignment = left_alignment
            ws_impact.row_dimensions[row_idx].height = 40

    # Ajustar larguras
    ws_impact.column_dimensions['A'].width = 30
    ws_impact.column_dimensions['B'].width = 35
    ws_impact.column_dimensions['C'].width = 40
    ws_impact.column_dimensions['D'].width = 35

    # ============================================================
    # ABA 4: ANÁLISE DE RISCOS (Detalhada)
    # ============================================================

    ws_riscos['A1'] = "ANÁLISE DETALHADA DE RISCOS"
    ws_riscos['A1'].font = Font(bold=True, size=14, color="1F4E78")
    ws_riscos.merge_cells('A1:P1')

    headers_riscos = [
        'ID', 'Ameaça', 'Atacante',
        'Skill', 'Motive', 'Opportunity', 'Size', 'THREAT SCORE',
        'Ease Disc.', 'Ease Exploit', 'Awareness', 'Intrusion Det.', 'VULN SCORE',
        'LIKELIHOOD', 'IMPACT', 'RISK SCORE', 'CRITICIDADE'
    ]

    for col, header in enumerate(headers_riscos, start=1):
        cell = ws_riscos.cell(row=2, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border
        ws_riscos.row_dimensions[2].height = 30

    # Dados das ameaças
    ameacas_data = [
        # ID, Ameaça, Atacante, Skill, Motive, Opportunity, Size, Ease_Disc, Ease_Exploit, Awareness, Intrusion_Det, Impact, Criticidade
        ('T01', 'Força Bruta sem Rate Limiting', 'Fraudador/BotHerder', 6, 9, 9, 9, 9, 9, 7, 9, 9, 'CRÍTICA'),
        ('T02', 'Automação sem CAPTCHA', 'BotHerder', 7, 8, 9, 9, 9, 9, 9, 9, 9, 'CRÍTICA'),
        ('T03', 'Ausência de Bloqueio Temporário', 'Fraudador', 5, 9, 9, 9, 7, 8, 7, 9, 9, 'CRÍTICA'),
        ('T04', 'Input Validation Inadequada', 'Fraudador', 6, 9, 7, 7, 7, 8, 7, 9, 9, 'CRÍTICA'),
        ('T05', 'Comunicação HTTP sem TLS', 'Qualquer', 4, 7, 5, 9, 6, 7, 9, 7, 8, 'CRÍTICA'),
        ('T06', 'Enumeração de Usuários', 'Fraudador/BotHerder', 5, 8, 9, 9, 8, 7, 7, 9, 7, 'ALTA'),
        ('T07', 'CSRF sem Proteção', 'Fraudador Oportunista', 5, 7, 6, 7, 6, 7, 7, 8, 8, 'ALTA'),
        ('T08', 'Ausência de Logging', 'Qualquer', 3, 6, 9, 9, 7, 7, 9, 9, 7, 'ALTA'),
        ('T09', 'Session Management Inseguro', 'Fraudador', 6, 8, 6, 7, 5, 6, 7, 7, 8, 'ALTA'),
        ('T10', 'Perguntas com Baixa Entropia', 'Fraudador/Carder', 4, 9, 8, 9, 9, 8, 9, 9, 8, 'ALTA'),
        ('T11', 'CORS Inadequado', 'Fraudador', 6, 5, 5, 7, 5, 6, 6, 7, 6, 'MÉDIA'),
        ('T12', 'Sem Notificação ao Usuário', 'Qualquer', 2, 4, 9, 9, 7, 7, 7, 9, 5, 'MÉDIA'),
        ('T13', 'Sem Device Fingerprinting', 'Fraudador', 5, 6, 7, 7, 6, 6, 5, 8, 6, 'MÉDIA'),
    ]

    for row_idx, (tid, ameaca, atacante, skill, motive, opp, size,
                  ease_disc, ease_expl, aware, intrusion, impact, crit) in enumerate(ameacas_data, start=3):

        # Calcular scores
        threat_score = round((skill + motive + opp + size) / 4, 1)
        vuln_score = round((ease_disc + ease_expl + aware + intrusion) / 4, 1)
        likelihood = round((threat_score + vuln_score) / 2, 1)
        risk_score = round((likelihood + impact) / 2, 1)

        # Preencher dados
        ws_riscos.cell(row=row_idx, column=1).value = tid
        ws_riscos.cell(row=row_idx, column=2).value = ameaca
        ws_riscos.cell(row=row_idx, column=3).value = atacante
        ws_riscos.cell(row=row_idx, column=4).value = skill
        ws_riscos.cell(row=row_idx, column=5).value = motive
        ws_riscos.cell(row=row_idx, column=6).value = opp
        ws_riscos.cell(row=row_idx, column=7).value = size
        ws_riscos.cell(row=row_idx, column=8).value = threat_score
        ws_riscos.cell(row=row_idx, column=9).value = ease_disc
        ws_riscos.cell(row=row_idx, column=10).value = ease_expl
        ws_riscos.cell(row=row_idx, column=11).value = aware
        ws_riscos.cell(row=row_idx, column=12).value = intrusion
        ws_riscos.cell(row=row_idx, column=13).value = vuln_score
        ws_riscos.cell(row=row_idx, column=14).value = likelihood
        ws_riscos.cell(row=row_idx, column=15).value = impact
        ws_riscos.cell(row=row_idx, column=16).value = risk_score
        ws_riscos.cell(row=row_idx, column=17).value = crit

        # Aplicar formatação
        for col in range(1, 18):
            cell = ws_riscos.cell(row=row_idx, column=col)
            cell.border = thin_border
            cell.alignment = center_alignment if col != 2 else left_alignment

            # Colorir criticidade
            if col == 17:
                if crit == 'CRÍTICA':
                    cell.fill = critico_fill
                    cell.font = critico_font
                elif crit == 'ALTA':
                    cell.fill = alto_fill
                    cell.font = alto_font
                elif crit == 'MÉDIA':
                    cell.fill = medio_fill
                    cell.font = medio_font

        ws_riscos.row_dimensions[row_idx].height = 30

    # Ajustar larguras
    widths = [8, 35, 25, 8, 8, 10, 8, 12, 10, 10, 10, 12, 12, 12, 10, 12, 14]
    for col, width in enumerate(widths, start=1):
        ws_riscos.column_dimensions[get_column_letter(col)].width = width

    # ============================================================
    # ABA 5: PLANO DE MITIGAÇÃO
    # ============================================================

    ws_mitigacoes['A1'] = "PLANO DE MITIGAÇÃO"
    ws_mitigacoes['A1'].font = Font(bold=True, size=14, color="1F4E78")
    ws_mitigacoes.merge_cells('A1:F1')

    headers_mitigacao = ['ID', 'Ameaça', 'Criticidade', 'Mitigação (BFF/Frontend)', 'Camada', 'Prioridade']

    for col, header in enumerate(headers_mitigacao, start=1):
        cell = ws_mitigacoes.cell(row=2, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border

    mitigacoes_data = [
        ('T01', 'Força Bruta sem Rate Limiting', 'CRÍTICA',
         'Rate limiting multi-camada: 3 tent./15min por CPF, 10 tent./hora por IP, 5 tent./sessão. Bloqueio progressivo: 15min → 1h → 24h',
         'BFF', 'P0 - Fase 1'),

        ('T02', 'Automação sem CAPTCHA', 'CRÍTICA',
         'reCAPTCHA v3 (score-based) invisível em todas páginas + reCAPTCHA v2 (desafio) após 2 falhas. Validação server-side no BFF',
         'BFF + Frontend', 'P0 - Fase 1'),

        ('T03', 'Ausência de Bloqueio Temporário', 'CRÍTICA',
         'Implementar bloqueio temporário progressivo: 3 falhas=15min, 5=1h, 10=24h, 20=definitivo. Contador persistente (Redis)',
         'BFF', 'P0 - Fase 1'),

        ('T04', 'Input Validation Inadequada', 'CRÍTICA',
         'Validação rigorosa de CPF (formato+dígitos), sanitização de respostas (remove chars perigosos), length limits (100 chars), whitelist',
         'BFF + Frontend', 'P0 - Fase 1'),

        ('T05', 'Comunicação HTTP sem TLS', 'CRÍTICA',
         'HTTPS obrigatório, HSTS header (maxAge: 1 ano, includeSubDomains), redirect HTTP→HTTPS, certificado SSL/TLS válido',
         'BFF + Frontend', 'P0 - Fase 1'),

        ('T06', 'Enumeração de Usuários', 'ALTA',
         'Resposta uniforme para CPF válido/inválido, sempre exibir formulário de perguntas, timing consistente (artificial delay)',
         'BFF', 'P1 - Fase 2'),

        ('T07', 'CSRF sem Proteção', 'ALTA',
         'Implementar tokens anti-CSRF (csurf middleware), gerar token em cada sessão, validar em todos POST/PUT/DELETE',
         'BFF + Frontend', 'P1 - Fase 2'),

        ('T08', 'Ausência de Logging', 'ALTA',
         'Logging estruturado (winston) de todas tentativas, alertas automáticos após 3 falhas, métricas em tempo real (Prometheus)',
         'BFF', 'P1 - Fase 2'),

        ('T09', 'Session Management Inseguro', 'ALTA',
         'express-session + Redis, cookies seguros (httpOnly, secure, sameSite:strict), timeout 15min, regeneração após eventos críticos',
         'BFF', 'P1 - Fase 2'),

        ('T10', 'Perguntas com Baixa Entropia', 'ALTA',
         'Perguntas de alta entropia: valor última transação, nome empresa atual, código 6 dígitos, data última senha, agência bancária',
         'BFF', 'P1 - Fase 2'),

        ('T11', 'CORS Inadequado', 'MÉDIA',
         'CORS restritivo: origin apenas domínio do Frontend, credentials:true, métodos limitados (GET,POST), headers específicos',
         'BFF', 'P2 - Fase 3'),

        ('T12', 'Sem Notificação ao Usuário', 'MÉDIA',
         'Enviar SMS/Email ao usuário legítimo após 3 tentativas falhadas e após recuperação bem-sucedida',
         'BFF', 'P2 - Fase 3'),

        ('T13', 'Sem Device Fingerprinting', 'MÉDIA',
         'Coletar fingerprint do dispositivo (FingerprintJS), exigir validação adicional em dispositivos desconhecidos',
         'BFF + Frontend', 'P2 - Fase 3'),
    ]

    for row_idx, (tid, ameaca, crit, mitigacao, camada, prioridade) in enumerate(mitigacoes_data, start=3):
        ws_mitigacoes.cell(row=row_idx, column=1).value = tid
        ws_mitigacoes.cell(row=row_idx, column=2).value = ameaca
        ws_mitigacoes.cell(row=row_idx, column=3).value = crit
        ws_mitigacoes.cell(row=row_idx, column=4).value = mitigacao
        ws_mitigacoes.cell(row=row_idx, column=5).value = camada
        ws_mitigacoes.cell(row=row_idx, column=6).value = prioridade

        for col in range(1, 7):
            cell = ws_mitigacoes.cell(row=row_idx, column=col)
            cell.border = thin_border
            cell.alignment = left_alignment

            # Colorir criticidade
            if col == 3:
                if crit == 'CRÍTICA':
                    cell.fill = critico_fill
                    cell.font = critico_font
                elif crit == 'ALTA':
                    cell.fill = alto_fill
                    cell.font = alto_font
                elif crit == 'MÉDIA':
                    cell.fill = medio_fill
                    cell.font = medio_font

        ws_mitigacoes.row_dimensions[row_idx].height = 50

    # Ajustar larguras
    ws_mitigacoes.column_dimensions['A'].width = 8
    ws_mitigacoes.column_dimensions['B'].width = 30
    ws_mitigacoes.column_dimensions['C'].width = 12
    ws_mitigacoes.column_dimensions['D'].width = 70
    ws_mitigacoes.column_dimensions['E'].width = 15
    ws_mitigacoes.column_dimensions['F'].width = 15

    # Salvar planilha
    wb.save('ETAPA2-OWASP-Risk-Rating.xlsx')
    print("✅ Planilha OWASP Risk Rating criada com sucesso: ETAPA2-OWASP-Risk-Rating.xlsx")

if __name__ == "__main__":
    criar_planilha_owasp()
